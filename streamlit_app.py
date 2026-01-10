import os
import io
from datetime import date

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from supabase import create_client, Client

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# =========================
# CONFIG
# =========================
st.set_page_config(page_title="Trip Logbook", layout="centered")

SUPABASE_URL = st.secrets.get("SUPABASE_URL", os.getenv("SUPABASE_URL"))
SUPABASE_SERVICE_ROLE_KEY = st.secrets.get("SUPABASE_SERVICE_ROLE_KEY", os.getenv("SUPABASE_SERVICE_ROLE_KEY"))
SUPABASE_KEY_FALLBACK = st.secrets.get("SUPABASE_KEY", os.getenv("SUPABASE_KEY", ""))
ADMIN_PIN = st.secrets.get("ADMIN_PIN", os.getenv("ADMIN_PIN", ""))
APP_TITLE = st.secrets.get("APP_TITLE", os.getenv("APP_TITLE", "🚗 Trip Logbook"))

if not SUPABASE_URL:
    st.error("Missing SUPABASE_URL in Streamlit secrets.")
    st.stop()

SUPABASE_KEY = SUPABASE_SERVICE_ROLE_KEY or SUPABASE_KEY_FALLBACK
if not SUPABASE_KEY:
    st.error("Missing SUPABASE_SERVICE_ROLE_KEY (recommended) or SUPABASE_KEY in secrets.")
    st.stop()

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

PLACE_TYPE_NEW = "✍️ Type a new place..."


# =========================
# BASIC HELPERS
# =========================
def show_api_error(e: Exception):
    st.error("Database request failed.")
    st.write("Details (for debugging):")
    st.exception(e)


def clean_text(s: str) -> str:
    if not s:
        return ""
    return " ".join(str(s).strip().split())


def canonical_place(s: str) -> str:
    return clean_text(s).lower()


def normalize_pair(a: str, b: str) -> tuple[str, str]:
    a = canonical_place(a)
    b = canonical_place(b)
    return (a, b) if a <= b else (b, a)


def parse_year_from_period_name(name: str) -> int:
    name = clean_text(name)
    if name.isdigit() and len(name) == 4:
        return int(name)
    return date.today().year


# =========================
# iOS-LIKE WHEEL PICKER (Distance)
# - compact by default
# - tap/click to open wheel
# =========================
def distance_wheel_picker_html(
    key: str,
    value: float,
    min_value: float = 1.0,
    max_value: float = 200.0,
    step: float = 0.5,
    height_px: int = 160,
) -> float:
    """
    iOS-style wheel picker using an HTML component.
    Returns float. Uses wrap effect by repeating list 3x.
    Compatibility: does NOT pass key= into components.html().
    """
    values = []
    steps = int(round((max_value - min_value) / step))
    for i in range(steps + 1):
        values.append(round(min_value + i * step, 1))

    if value is None:
        value = min_value
    value = float(value)
    value = max(min_value, min(max_value, value))
    value = round(value / step) * step
    value = round(value, 1)

    values_3x = values + values + values
    base_len = len(values)

    try:
        idx_in_base = values.index(value)
    except ValueError:
        idx_in_base = 0
    initial_index = base_len + idx_in_base

    options_html = "\n".join(
        f'<div class="item" data-value="{val:.1f}">{val:.1f}</div>'
        for val in values_3x
    )

    list_id = f"list_{key}"

    html = f"""
    <html>
      <head>
        <meta name="viewport" content="width=device-width, initial-scale=1.0" />
        <script src="https://unpkg.com/streamlit-component-lib@1.6.0/dist/streamlit-component-lib.js"></script>
        <style>
          :root {{
            --h: 28px;  /* row height */
          }}
          body {{
            margin: 0;
            padding: 0;
            background: transparent;
            color: #eaeaea;
            font-family: system-ui, -apple-system, Segoe UI, Roboto, sans-serif;
          }}
          .wheel {{
            position: relative;
            height: {height_px}px;
            border-radius: 12px;
            border: 1px solid rgba(255,255,255,0.15);
            background: rgba(255,255,255,0.04);
            overflow: hidden;
          }}
          .list {{
            height: 100%;
            overflow-y: scroll;
            scrollbar-width: thin;
            scroll-snap-type: y mandatory;
            padding: calc({height_px}px/2 - var(--h)/2) 0;
          }}
          .item {{
            height: var(--h);
            line-height: var(--h);
            text-align: center;
            font-size: 15px;
            scroll-snap-align: center;
            color: rgba(234,234,234,0.70);
            user-select: none;
          }}
          .item.active {{
            color: rgba(234,234,234,1.0);
            font-weight: 800;
            font-size: 17px;
          }}
          .center-highlight {{
            position: absolute;
            left: 10px;
            right: 10px;
            top: 50%;
            transform: translateY(-50%);
            height: var(--h);
            border-radius: 10px;
            background: rgba(255,255,255,0.10);
            border: 1px solid rgba(255,255,255,0.18);
            pointer-events: none;
          }}
          .fade-top, .fade-bottom {{
            position: absolute;
            left: 0;
            right: 0;
            height: 36px;
            pointer-events: none;
          }}
          .fade-top {{
            top: 0;
            background: linear-gradient(to bottom, rgba(18,18,18,0.90), rgba(18,18,18,0));
          }}
          .fade-bottom {{
            bottom: 0;
            background: linear-gradient(to top, rgba(18,18,18,0.90), rgba(18,18,18,0));
          }}
        </style>
      </head>
      <body>
        <div class="wheel">
          <div class="fade-top"></div>
          <div class="fade-bottom"></div>
          <div class="center-highlight"></div>
          <div id="{list_id}" class="list">
            {options_html}
          </div>
        </div>

        <script>
          const list = document.getElementById("{list_id}");
          const rowH = parseFloat(getComputedStyle(document.documentElement).getPropertyValue('--h'));
          const baseLen = {base_len};
          const initialIndex = {initial_index};

          function setActiveByIndex(idx) {{
            const items = list.querySelectorAll(".item");
            items.forEach(el => el.classList.remove("active"));
            if (items[idx]) items[idx].classList.add("active");
          }}

          function nearestIndex() {{
            const center = list.scrollTop + (list.clientHeight / 2);
            const idx = Math.round((center - (rowH / 2)) / rowH);
            return Math.max(0, Math.min(idx, list.children.length - 1));
          }}

          function valueAtIndex(idx) {{
            const el = list.children[idx];
            if (!el) return null;
            return parseFloat(el.dataset.value);
          }}

          let debounce = null;

          function sendValue(val) {{
            if (window.Streamlit) {{
              window.Streamlit.setComponentValue(val);
            }}
          }}

          function snapAndWrap() {{
            const idx = nearestIndex();
            const val = valueAtIndex(idx);
            if (val === null) return;

            const targetScroll = idx * rowH - (list.clientHeight/2 - rowH/2);
            list.scrollTo({{ top: targetScroll, behavior: "auto" }});

            let idxIn3 = idx;
            if (idxIn3 < baseLen) {{
              idxIn3 = idxIn3 + baseLen;
            }} else if (idxIn3 >= baseLen*2) {{
              idxIn3 = idxIn3 - baseLen;
            }}

            const val2 = valueAtIndex(idxIn3);
            const targetScroll2 = idxIn3 * rowH - (list.clientHeight/2 - rowH/2);
            if (Math.abs(targetScroll2 - list.scrollTop) > 1) {{
              list.scrollTo({{ top: targetScroll2, behavior: "auto" }});
            }}

            setActiveByIndex(idxIn3);
            sendValue(val2);
          }}

          function init() {{
            const targetScroll = initialIndex * rowH - (list.clientHeight/2 - rowH/2);
            list.scrollTo({{ top: targetScroll, behavior: "auto" }});
            setActiveByIndex(initialIndex);
            sendValue(valueAtIndex(initialIndex));
            if (window.Streamlit) window.Streamlit.setFrameHeight({height_px});
          }}

          list.addEventListener("scroll", () => {{
            if (debounce) clearTimeout(debounce);
            debounce = setTimeout(snapAndWrap, 140);
          }});

          list.addEventListener("click", (e) => {{
            const item = e.target.closest(".item");
            if (!item) return;
            const items = Array.from(list.children);
            const idx = items.indexOf(item);
            const targetScroll = idx * rowH - (list.clientHeight/2 - rowH/2);
            list.scrollTo({{ top: targetScroll, behavior: "smooth" }});
            if (debounce) clearTimeout(debounce);
            debounce = setTimeout(snapAndWrap, 180);
          }});

          window.addEventListener("load", init);
        </script>
      </body>
    </html>
    """

    picked = components.html(html, height=height_px)
    if picked is None:
        return value
    try:
        return float(picked)
    except Exception:
        return value


# =========================
# SUPABASE HELPERS
# =========================
def get_periods():
    try:
        res = (
            supabase.table("periods")
            .select("id,name,is_active")
            .eq("is_active", True)
            .order("name", desc=True)
            .execute()
        )
        return res.data or []
    except Exception as e:
        show_api_error(e)
        return []


def ensure_period(name: str):
    name = clean_text(name)
    if not name:
        return None
    try:
        existing = supabase.table("periods").select("id").eq("name", name).limit(1).execute().data
        if existing:
            return existing[0]["id"]
        created = supabase.table("periods").insert({"name": name, "is_active": True}).execute().data
        if created:
            return created[0]["id"]
    except Exception as e:
        show_api_error(e)
    return None


def create_period(name: str):
    name = clean_text(name)
    if not name:
        return None
    try:
        existing = supabase.table("periods").select("id,name").eq("name", name).limit(1).execute().data
        if existing:
            return existing[0]
        res = supabase.table("periods").insert({"name": name, "is_active": True}).execute()
        rows = res.data or []
        return rows[0] if rows else None
    except Exception as e:
        show_api_error(e)
        return None


def get_cars():
    try:
        res = (
            supabase.table("cars")
            .select("id,name,plate,is_active")
            .eq("is_active", True)
            .order("name")
            .execute()
        )
        return res.data or []
    except Exception as e:
        show_api_error(e)
        return []


def get_places(limit=1000):
    try:
        res = (
            supabase.table("places")
            .select("name,is_active")
            .eq("is_active", True)
            .order("name")
            .limit(limit)
            .execute()
        )
        rows = res.data or []
        return [r["name"] for r in rows if r.get("name")]
    except Exception as e:
        show_api_error(e)
        return []


def upsert_place(name: str):
    name = clean_text(name)
    if not name:
        return
    try:
        existing = supabase.table("places").select("id").eq("name", name).limit(1).execute().data
        if existing:
            pid = existing[0]["id"]
            supabase.table("places").update({"is_active": True}).eq("id", pid).execute()
        else:
            supabase.table("places").insert({"name": name, "is_active": True}).execute()
    except Exception:
        pass


def get_route_distance(departure: str, arrival: str) -> float | None:
    dep = canonical_place(departure)
    arr = canonical_place(arrival)
    if not dep or not arr:
        return None
    a, b = normalize_pair(dep, arr)
    try:
        res = (
            supabase.table("route_distances")
            .select("distance_km")
            .eq("place_a", a)
            .eq("place_b", b)
            .limit(1)
            .execute()
        )
        rows = res.data or []
        if not rows:
            return None
        return float(rows[0]["distance_km"])
    except Exception:
        return None


def set_route_distance(departure: str, arrival: str, distance_km: float):
    dep = canonical_place(departure)
    arr = canonical_place(arrival)
    if not dep or not arr:
        return
    a, b = normalize_pair(dep, arr)
    try:
        existing = (
            supabase.table("route_distances")
            .select("id")
            .eq("place_a", a)
            .eq("place_b", b)
            .limit(1)
            .execute()
        ).data

        if existing:
            rid = existing[0]["id"]
            supabase.table("route_distances").update({"distance_km": float(distance_km)}).eq("id", rid).execute()
        else:
            supabase.table("route_distances").insert(
                {"place_a": a, "place_b": b, "distance_km": float(distance_km)}
            ).execute()
    except Exception:
        pass


def insert_trip(period_id: str, trip_date: date, car_id: str, departure: str, arrival: str, distance_km: float, notes: str):
    payload = {
        "period_id": period_id,
        "trip_date": str(trip_date),
        "car_id": car_id,
        "departure": clean_text(departure),
        "arrival": clean_text(arrival),
        "distance_km": float(distance_km),
        "notes": clean_text(notes) if notes else None,
    }
    try:
        return supabase.table("trip_entries").insert(payload).execute()
    except Exception as e:
        show_api_error(e)
        return None


def fetch_entries(period_id: str, start_date: date, end_date: date, car_id=None, search_text=""):
    try:
        q = (
            supabase.table("trip_entries")
            .select("id,period_id,trip_date,car_id,departure,arrival,distance_km,notes,created_at")
            .eq("period_id", period_id)
            .gte("trip_date", str(start_date))
            .lte("trip_date", str(end_date))
            .order("trip_date", desc=False)
            .order("created_at", desc=False)
        )
        if car_id:
            q = q.eq("car_id", car_id)

        res = q.execute()
        rows = res.data or []
        df = pd.DataFrame(rows)

        if df.empty:
            return df

        if search_text.strip():
            s = search_text.strip().lower()
            mask = (
                df["departure"].astype(str).str.lower().str.contains(s, na=False)
                | df["arrival"].astype(str).str.lower().str.contains(s, na=False)
                | df["notes"].astype(str).str.lower().str.contains(s, na=False)
            )
            df = df[mask].copy()

        return df
    except Exception as e:
        show_api_error(e)
        return pd.DataFrame([])


def update_trip(trip_id: str, updates: dict):
    allowed = {"trip_date", "car_id", "departure", "arrival", "distance_km", "notes"}
    updates_clean = {k: v for k, v in updates.items() if k in allowed}
    try:
        return supabase.table("trip_entries").update(updates_clean).eq("id", trip_id).execute()
    except Exception as e:
        show_api_error(e)
        return None


def delete_trips(trip_ids: list[str]):
    if not trip_ids:
        return
    try:
        return supabase.table("trip_entries").delete().in_("id", trip_ids).execute()
    except Exception as e:
        show_api_error(e)
        return None


def fetch_places_admin():
    try:
        res = supabase.table("places").select("id,name,is_active,created_at").order("name").execute()
        return pd.DataFrame(res.data or [])
    except Exception as e:
        show_api_error(e)
        return pd.DataFrame([])


def update_place(place_id: str, updates: dict):
    allowed = {"name", "is_active"}
    updates_clean = {k: v for k, v in updates.items() if k in allowed}
    try:
        return supabase.table("places").update(updates_clean).eq("id", place_id).execute()
    except Exception as e:
        show_api_error(e)
        return None


# =========================
# RANGE + GROUPING HELPERS
# =========================
def month_range(d: date):
    start = d.replace(day=1)
    if start.month == 12:
        next_month = date(start.year + 1, 1, 1)
    else:
        next_month = date(start.year, start.month + 1, 1)
    last_day = (pd.Timestamp(next_month) - pd.Timedelta(days=1)).date()
    return start, last_day


def year_range(year: int):
    return date(year, 1, 1), date(year, 12, 31)


def add_month_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out["trip_date"] = pd.to_datetime(out["trip_date"], errors="coerce")
    out["month_key"] = out["trip_date"].dt.to_period("M").astype(str)
    out["month_name"] = out["trip_date"].dt.strftime("%B %Y")
    return out


def monthly_totals(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Month", "Total distance (km)"])
    d = add_month_columns(df)
    d["distance_km"] = pd.to_numeric(d["distance_km"], errors="coerce").fillna(0.0)
    totals = (
        d.groupby(["month_key", "month_name"], as_index=False)["distance_km"]
        .sum()
        .sort_values("month_key")
    )
    totals = totals.rename(columns={"month_name": "Month", "distance_km": "Total distance (km)"})
    totals["Total distance (km)"] = totals["Total distance (km)"].round(1)
    return totals[["Month", "Total distance (km)"]]


# =========================
# EXPORT PREP
# =========================
def make_export_df(df: pd.DataFrame, car_id_to_label: dict) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out["Car"] = out["car_id"].map(car_id_to_label).fillna("")
    out["Date"] = pd.to_datetime(out["trip_date"]).dt.date.astype(str)
    out = out.rename(columns={
        "departure": "Departure",
        "arrival": "Arrival",
        "distance_km": "Distance (km)",
        "notes": "Notes",
    })
    out["Distance (km)"] = pd.to_numeric(out["Distance (km)"], errors="coerce").fillna(0.0).round(1)
    out = out[["Date", "Car", "Departure", "Arrival", "Distance (km)", "Notes"]]
    out["Notes"] = out["Notes"].fillna("")
    return out


def export_csv_bytes(df: pd.DataFrame) -> bytes:
    buff = io.StringIO()
    df.to_csv(buff, index=False)
    return buff.getvalue().encode("utf-8")


def _autosize_worksheet(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def export_xlsx_bytes_grouped(df_export: pd.DataFrame, df_raw: pd.DataFrame, title: str) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary["A1"] = title
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"] = "Monthly totals"
    ws_summary["A2"].font = Font(bold=True, size=12)

    totals = monthly_totals(df_raw)
    if totals.empty:
        ws_summary["A4"] = "No trips in this selection."
    else:
        start_row = 4
        for r_idx, row in enumerate(dataframe_to_rows(totals, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    ws_summary.cell(row=r_idx, column=c_idx).font = Font(bold=True)
        ws_summary.freeze_panes = "A5"
        _autosize_worksheet(ws_summary)

    if not df_raw.empty and not df_export.empty:
        d = add_month_columns(df_raw)
        work = df_export.copy()
        work["month_key"] = d["month_key"].values
        work["month_name"] = d["month_name"].values

        month_keys = sorted(work["month_key"].dropna().unique().tolist())
        for mk in month_keys:
            block = work[work["month_key"] == mk].copy()
            month_name = block["month_name"].iloc[0] if not block.empty else mk

            dt = pd.Period(mk).to_timestamp()
            sheet_name = dt.strftime("%b")
            if sheet_name in wb.sheetnames:
                n = 2
                while f"{sheet_name}{n}" in wb.sheetnames:
                    n += 1
                sheet_name = f"{sheet_name}{n}"

            ws = wb.create_sheet(title=sheet_name)
            ws["A1"] = month_name
            ws["A1"].font = Font(bold=True, size=13)
            ws["A2"] = f"Monthly total: {block['Distance (km)'].sum():.1f} km"
            ws["A2"].font = Font(bold=True)

            table = block[["Date", "Car", "Departure", "Arrival", "Distance (km)", "Notes"]].copy()
            start_row = 4
            for r_idx, row in enumerate(dataframe_to_rows(table, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row:
                        ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)

            ws.freeze_panes = "A5"
            _autosize_worksheet(ws)

    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def export_pdf_bytes_grouped(df_export: pd.DataFrame, df_raw: pd.DataFrame, title: str) -> bytes:
    buff = io.BytesIO()
    c = canvas.Canvas(buff, pagesize=A4)

    width, height = A4
    left = 40
    top = height - 50

    def new_page():
        c.showPage()
        return height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(left, top, title)
    y = top - 24

    totals = monthly_totals(df_raw)
    c.setFont("Helvetica", 11)
    if totals.empty:
        c.drawString(left, y, "No trips in this selection.")
        c.showPage()
        c.save()
        return buff.getvalue()

    overall_total = float(pd.to_numeric(df_raw["distance_km"], errors="coerce").fillna(0.0).sum())
    c.drawString(left, y, f"Total distance: {overall_total:.1f} km")
    y = new_page()

    d = add_month_columns(df_raw)
    work = df_export.copy()
    work["month_key"] = d["month_key"].values
    work["month_name"] = d["month_name"].values

    month_keys = sorted(work["month_key"].dropna().unique().tolist())
    for mk in month_keys:
        block = work[work["month_key"] == mk].copy()
        month_name = block["month_name"].iloc[0]
        month_total = float(block["Distance (km)"].sum())

        c.setFont("Helvetica-Bold", 14)
        c.drawString(left, y, month_name)
        y -= 16

        c.setFont("Helvetica-Bold", 11)
        c.drawString(left, y, f"Monthly total: {month_total:.1f} km")
        y -= 18

        headers = ["Date", "Car", "Departure", "Arrival", "Km", "Notes"]
        col_widths = [70, 70, 120, 120, 40, 110]

        c.setFont("Helvetica-Bold", 9)
        xx = left
        for h, cw in zip(headers, col_widths):
            c.drawString(xx, y, h)
            xx += cw
        y -= 12

        c.setFont("Helvetica", 9)
        for _, row in block.iterrows():
            if y < 70:
                y = new_page()
                c.setFont("Helvetica-Bold", 14)
                c.drawString(left, y, month_name + " (cont.)")
                y -= 16
                c.setFont("Helvetica-Bold", 11)
                c.drawString(left, y, f"Monthly total: {month_total:.1f} km")
                y -= 18
                c.setFont("Helvetica-Bold", 9)
                xx = left
                for h, cw in zip(headers, col_widths):
                    c.drawString(xx, y, h)
                    xx += cw
                y -= 12
                c.setFont("Helvetica", 9)

            values = [
                str(row.get("Date", ""))[:12],
                str(row.get("Car", ""))[:12],
                str(row.get("Departure", ""))[:22],
                str(row.get("Arrival", ""))[:22],
                f"{float(row.get('Distance (km)', 0.0)):.1f}",
                str(row.get("Notes", ""))[:22],
            ]
            xx = left
            for v, cw in zip(values, col_widths):
                c.drawString(xx, y, v)
                xx += cw
            y -= 12

        y -= 14
        if mk != month_keys[-1]:
            y = new_page()

    c.showPage()
    c.save()
    return buff.getvalue()


# =========================
# SESSION STATE
# =========================
if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

if "dep_choice" not in st.session_state:
    st.session_state.dep_choice = ""
if "arr_choice" not in st.session_state:
    st.session_state.arr_choice = ""
if "dep_typed" not in st.session_state:
    st.session_state.dep_typed = ""
if "arr_typed" not in st.session_state:
    st.session_state.arr_typed = ""

if "distance_value" not in st.session_state:
    st.session_state.distance_value = 1.0
if "distance_manual" not in st.session_state:
    st.session_state.distance_manual = False
if "last_route_key" not in st.session_state:
    st.session_state.last_route_key = ""


def get_dep_value(places_list: list[str]) -> str:
    if st.session_state.dep_choice == PLACE_TYPE_NEW:
        return clean_text(st.session_state.dep_typed)
    if st.session_state.dep_choice in places_list:
        return clean_text(st.session_state.dep_choice)
    return clean_text(st.session_state.dep_typed)


def get_arr_value(places_list: list[str]) -> str:
    if st.session_state.arr_choice == PLACE_TYPE_NEW:
        return clean_text(st.session_state.arr_typed)
    if st.session_state.arr_choice in places_list:
        return clean_text(st.session_state.arr_choice)
    return clean_text(st.session_state.arr_typed)


def route_key_for_current(places_list: list[str]) -> str:
    dep = canonical_place(get_dep_value(places_list))
    arr = canonical_place(get_arr_value(places_list))
    if not dep or not arr:
        return ""
    a, b = normalize_pair(dep, arr)
    return f"{a}__{b}"


def maybe_autofill_distance(places_list: list[str]):
    rk = route_key_for_current(places_list)
    if not rk:
        return
    if rk != st.session_state.last_route_key:
        st.session_state.last_route_key = rk
        st.session_state.distance_manual = False
        dep = get_dep_value(places_list)
        arr = get_arr_value(places_list)
        mem = get_route_distance(dep, arr)
        if mem is not None:
            mem = float(mem)
            mem = max(1.0, min(200.0, mem))
            mem = round(mem * 2) / 2.0  # nearest 0.5
            st.session_state.distance_value = mem


# =========================
# MAIN UI
# =========================
st.title(APP_TITLE)
tabs = st.tabs(["🧾 Trip Log", "🛠️ Admin"])


# ---------- TRIP LOG TAB ----------
with tabs[0]:
    # Admin unlock
    with st.expander("🔐 Admin mode"):
        if not ADMIN_PIN:
            st.info("Set ADMIN_PIN in Streamlit secrets to enable Admin features.")
        pin = st.text_input("Enter admin PIN", type="password")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Unlock admin", use_container_width=True):
                if ADMIN_PIN and pin == ADMIN_PIN:
                    st.session_state.is_admin = True
                    st.success("Admin mode enabled.")
                else:
                    st.error("Wrong PIN.")
        with c2:
            if st.button("Lock admin", use_container_width=True):
                st.session_state.is_admin = False
                st.info("Admin mode disabled.")

    # Period selection (default current year)
    st.subheader("Period (logbook)")
    current_year_name = str(date.today().year)
    ensure_period(current_year_name)

    periods = get_periods()
    if not periods:
        ensure_period("Default")
        periods = get_periods()

    period_names = [p["name"] for p in periods]
    period_name_to_id = {p["name"]: p["id"] for p in periods}
    default_index = period_names.index(current_year_name) if current_year_name in period_names else 0

    colp1, colp2 = st.columns([2, 1])
    with colp1:
        selected_period_name = st.selectbox("Choose period", period_names, index=default_index)
        selected_period_id = period_name_to_id[selected_period_name]
    with colp2:
        new_period_name = st.text_input("New period name", placeholder="e.g. 2027")
        if st.button("➕ Add period", use_container_width=True):
            if new_period_name.strip():
                create_period(new_period_name.strip())
                st.success("Period added.")
                st.rerun()
            else:
                st.error("Type a name first (e.g. 2027).")

    # Cars
    cars = get_cars()
    if not cars:
        st.error("No active cars found. Add your cars in the 'cars' table.")
        st.stop()

    car_label_to_id = {}
    car_id_to_label = {}
    for c in cars:
        label = c["name"] + (f" ({c['plate']})" if c.get("plate") else "")
        car_label_to_id[label] = c["id"]
        car_id_to_label[c["id"]] = label
    car_labels = list(car_label_to_id.keys())

    # Places
    places = get_places()
    dep_options = places + [PLACE_TYPE_NEW]
    arr_options = places + [PLACE_TYPE_NEW]

    if st.session_state.dep_choice and st.session_state.dep_choice not in dep_options:
        st.session_state.dep_choice = PLACE_TYPE_NEW
    if st.session_state.arr_choice and st.session_state.arr_choice not in arr_options:
        st.session_state.arr_choice = PLACE_TYPE_NEW

    # ----- Add trip -----
    st.subheader("Add a trip")
    with st.container(border=True):
        col1, col2 = st.columns(2)
        with col1:
            trip_date = st.date_input("Date", value=date.today())
            st.caption(f"Day: **{trip_date.strftime('%A')}**")
        with col2:
            car_label = st.selectbox("Car", car_labels)
            car_id = car_label_to_id[car_label]

        swap_col1, swap_col2 = st.columns([1, 3])
        with swap_col1:
            if st.button("↔ Swap", use_container_width=True):
                st.session_state.dep_choice, st.session_state.arr_choice = (
                    st.session_state.arr_choice,
                    st.session_state.dep_choice,
                )
                st.session_state.dep_typed, st.session_state.arr_typed = (
                    st.session_state.arr_typed,
                    st.session_state.dep_typed,
                )
                maybe_autofill_distance(places)
                st.rerun()
        with swap_col2:
            st.caption("Swap Departure and Arrival")

        st.selectbox(
            "Departure",
            dep_options,
            index=dep_options.index(st.session_state.dep_choice) if st.session_state.dep_choice in dep_options else 0,
            key="dep_choice",
            on_change=lambda: maybe_autofill_distance(places),
        )
        if st.session_state.dep_choice == PLACE_TYPE_NEW:
            st.text_input(
                "Departure (type)",
                value=st.session_state.dep_typed,
                key="dep_typed",
                on_change=lambda: maybe_autofill_distance(places),
            )

        st.selectbox(
            "Arrival",
            arr_options,
            index=arr_options.index(st.session_state.arr_choice) if st.session_state.arr_choice in arr_options else 0,
            key="arr_choice",
            on_change=lambda: maybe_autofill_distance(places),
        )
        if st.session_state.arr_choice == PLACE_TYPE_NEW:
            st.text_input(
                "Arrival (type)",
                value=st.session_state.arr_typed,
                key="arr_typed",
                on_change=lambda: maybe_autofill_distance(places),
            )

        maybe_autofill_distance(places)

        col3, col4 = st.columns(2)
        with col3:
            # Compact "input field" look + tap to open picker
            st.caption("Distance (km)")
            st.markdown(
                f"""
                <div style="
                    display:flex;
                    align-items:center;
                    justify-content:space-between;
                    padding:10px 12px;
                    border-radius:12px;
                    border:1px solid rgba(255,255,255,0.15);
                    background: rgba(255,255,255,0.04);
                    ">
                    <div style="font-size:18px; font-weight:800;">
                        {float(st.session_state.distance_value):.1f} km
                    </div>
                    <div style="opacity:0.7; font-size:14px;">
                        ✏️ edit
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

            has_popover = hasattr(st, "popover")
            if has_popover:
                with st.popover("Change distance", use_container_width=True):
                    picked = distance_wheel_picker_html(
                        key="distance_wheel",
                        value=float(st.session_state.distance_value),
                        min_value=1.0,
                        max_value=200.0,
                        step=0.5,
                        height_px=160,  # smaller wheel
                    )
                    if float(picked) != float(st.session_state.distance_value):
                        st.session_state.distance_value = float(picked)
                        st.session_state.distance_manual = True
                    st.caption("Scroll to select. Tap outside to close.")
            else:
                with st.expander("Change distance"):
                    picked = distance_wheel_picker_html(
                        key="distance_wheel",
                        value=float(st.session_state.distance_value),
                        min_value=1.0,
                        max_value=200.0,
                        step=0.5,
                        height_px=160,
                    )
                    if float(picked) != float(st.session_state.distance_value):
                        st.session_state.distance_value = float(picked)
                        st.session_state.distance_manual = True
                    st.caption("Scroll to select. Close this section when done.")

            distance = float(st.session_state.distance_value)

        with col4:
            notes = st.text_input("Notes (optional)")

        if st.button("✅ Save trip", use_container_width=True):
            departure = get_dep_value(places)
            arrival = get_arr_value(places)

            if not departure or not arrival:
                st.error("Please fill in Departure and Arrival.")
            else:
                insert_trip(selected_period_id, trip_date, car_id, departure, arrival, float(distance), notes)
                upsert_place(departure)
                upsert_place(arrival)
                set_route_distance(departure, arrival, float(distance))
                st.session_state.distance_manual = False
                st.success("Saved!")
                st.rerun()

    st.divider()

    # ----- View mode -----
    st.subheader("View trips")
    view_mode = st.radio(
        "View mode",
        ["One month", "All months (year)", "Custom range"],
        horizontal=True
    )

    year_for_period = parse_year_from_period_name(selected_period_name)

    if view_mode == "One month":
        pick = st.date_input("Pick a day in the month", value=date.today())
        start_date, end_date = month_range(pick)
    elif view_mode == "All months (year)":
        start_date, end_date = year_range(year_for_period)
        st.caption(f"Showing: **{year_for_period} (January → December)**")
    else:
        cA, cB = st.columns(2)
        with cA:
            start_date = st.date_input("Start", value=date(year_for_period, 1, 1))
        with cB:
            end_date = st.date_input("End", value=date.today())

    colf1, colf2 = st.columns(2)
    with colf1:
        filter_car = st.selectbox("Car filter", ["All cars"] + car_labels)
    with colf2:
        search_text = st.text_input("Search (departure/arrival/notes)", placeholder="type to search...")

    filter_car_id = None if filter_car == "All cars" else car_label_to_id[filter_car]

    df = fetch_entries(selected_period_id, start_date, end_date, car_id=filter_car_id, search_text=search_text)
    total_km = 0.0 if df.empty else float(pd.to_numeric(df["distance_km"], errors="coerce").fillna(0).sum())
    st.metric("Total distance (selected)", f"{total_km:.1f} km")

    if view_mode == "All months (year)" and not df.empty:
        st.write("### Total distance per month")
        st.dataframe(monthly_totals(df), use_container_width=True, hide_index=True)

    st.write("### Trips")
    df_export = make_export_df(df, car_id_to_label)

    if df.empty:
        st.info("No trips found.")
    else:
        if view_mode == "All months (year)":
            d = add_month_columns(df)
            df_export_block = df_export.copy()
            df_export_block["month_key"] = d["month_key"].values
            df_export_block["month_name"] = d["month_name"].values

            for mk in sorted(df_export_block["month_key"].unique().tolist()):
                block = df_export_block[df_export_block["month_key"] == mk].copy()
                month_name = block["month_name"].iloc[0]
                month_total = float(block["Distance (km)"].sum())

                st.subheader(month_name)
                st.caption(f"Monthly total: **{month_total:.1f} km**")
                st.dataframe(
                    block[["Date", "Car", "Departure", "Arrival", "Distance (km)", "Notes"]],
                    use_container_width=True,
                    hide_index=True,
                )
        else:
            st.dataframe(df_export, use_container_width=True, hide_index=True)

    st.divider()

    # Manage trips
    st.subheader("Manage trips (edit / delete)")
    if df.empty:
        st.info("Nothing to manage for this selection.")
    else:
        manage_df = df.copy()
        manage_df["trip_date"] = pd.to_datetime(manage_df["trip_date"], errors="coerce").dt.date
        manage_df["distance_km"] = pd.to_numeric(manage_df["distance_km"], errors="coerce").fillna(0.0).astype(float)
        manage_df["car_label"] = manage_df["car_id"].map(car_id_to_label).fillna("")
        manage_df["DELETE"] = False

        manage_df = manage_df[[
            "DELETE", "trip_date", "car_label", "departure", "arrival", "distance_km", "notes", "id"
        ]].copy()

        for col in ["departure", "arrival", "notes", "car_label", "id"]:
            manage_df[col] = manage_df[col].fillna("").astype(str)

        edited = st.data_editor(
            manage_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config={
                "DELETE": st.column_config.CheckboxColumn("Delete?"),
                "trip_date": st.column_config.DateColumn("Date"),
                "car_label": st.column_config.SelectboxColumn("Car", options=car_labels),
                "distance_km": st.column_config.NumberColumn("Distance (km)", min_value=0.0, step=0.1),
                "notes": st.column_config.TextColumn("Notes"),
                "id": st.column_config.TextColumn("ID", disabled=True),
            },
            disabled=["id"],
            key="manage_editor",
        )

        b1, b2 = st.columns(2)
        with b1:
            if st.button("💾 Save edits", use_container_width=True):
                changes = 0
                for i in range(len(edited)):
                    new_row = edited.iloc[i]
                    old_row = manage_df.iloc[i]
                    trip_id = str(new_row["id"]).strip()

                    if bool(new_row["DELETE"]):
                        continue

                    updates = {}

                    if str(new_row["trip_date"]) != str(old_row["trip_date"]):
                        updates["trip_date"] = str(pd.to_datetime(new_row["trip_date"]).date())

                    if str(new_row["car_label"]) != str(old_row["car_label"]):
                        updates["car_id"] = car_label_to_id.get(str(new_row["car_label"]))

                    for field in ["departure", "arrival", "notes"]:
                        nv = clean_text(new_row[field])
                        ov = clean_text(old_row[field])
                        if nv != ov:
                            updates[field] = nv

                    if float(new_row["distance_km"]) != float(old_row["distance_km"]):
                        updates["distance_km"] = float(new_row["distance_km"])

                    if updates:
                        update_trip(trip_id, updates)

                        dep_now = updates.get("departure", str(old_row["departure"]))
                        arr_now = updates.get("arrival", str(old_row["arrival"]))
                        dist_now = updates.get("distance_km", float(old_row["distance_km"]))

                        upsert_place(dep_now)
                        upsert_place(arr_now)
                        set_route_distance(dep_now, arr_now, float(dist_now))

                        changes += 1

                st.success(f"Saved edits on {changes} trip(s).")
                st.rerun()

        with b2:
            if st.button("🗑️ Delete selected", use_container_width=True):
                to_delete = edited.loc[edited["DELETE"] == True, "id"].astype(str).tolist()
                if not to_delete:
                    st.info("No trips selected.")
                else:
                    delete_trips(to_delete)
                    st.success(f"Deleted {len(to_delete)} trip(s).")
                    st.rerun()

    # Export at bottom
    st.divider()
    st.subheader("Export")

    default_name = f"{selected_period_name}_{start_date}_to_{end_date}"
    file_base = st.text_input("File name", value=default_name, help="Used for CSV / XLSX / PDF.")
    file_base = (file_base or "trips").strip().replace("/", "-")

    export_df = df_export if not df_export.empty else pd.DataFrame(
        columns=["Date", "Car", "Departure", "Arrival", "Distance (km)", "Notes"]
    )

    csv_bytes = export_csv_bytes(export_df)
    title = f"Trip Logbook — {selected_period_name}"
    xlsx_bytes = export_xlsx_bytes_grouped(export_df, df, title)
    pdf_bytes = export_pdf_bytes_grouped(export_df, df, title)

    e1, e2, e3 = st.columns(3)
    with e1:
        st.download_button("⬇️ CSV", csv_bytes, f"{file_base}.csv", "text/csv", use_container_width=True)
    with e2:
        st.download_button(
            "⬇️ XLSX",
            xlsx_bytes,
            f"{file_base}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with e3:
        st.download_button("⬇️ PDF", pdf_bytes, f"{file_base}.pdf", "application/pdf", use_container_width=True)


# ---------- ADMIN TAB ----------
with tabs[1]:
    if not st.session_state.is_admin:
        st.info("Admin is locked. Unlock it in the Trip Log tab.")
    else:
        st.header("Admin Panel")
        st.write("### Places manager")

        add_col1, add_col2 = st.columns([3, 1])
        with add_col1:
            new_place = st.text_input("Add a new place", placeholder="Type a place name (e.g. Amsterdam)")
        with add_col2:
            if st.button("➕ Add", use_container_width=True):
                if clean_text(new_place):
                    upsert_place(new_place)
                    st.success("Place added.")
                    st.rerun()
                else:
                    st.error("Type a place name first.")

        place_filter = st.text_input("Search places", placeholder="type to filter...")
        places_df = fetch_places_admin()

        if places_df.empty:
            st.info("No places yet. They appear automatically when trips are saved.")
        else:
            places_df = places_df[["id", "name", "is_active", "created_at"]].copy()
            places_df["name"] = places_df["name"].fillna("").astype(str)

            if place_filter.strip():
                s = place_filter.strip().lower()
                places_df = places_df[places_df["name"].str.lower().str.contains(s, na=False)].copy()

            st.caption("Edit names, or deactivate places (they won’t show in the dropdown).")
            edited_places = st.data_editor(
                places_df,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                column_config={
                    "id": st.column_config.TextColumn("ID", disabled=True),
                    "name": st.column_config.TextColumn("Place name"),
                    "is_active": st.column_config.CheckboxColumn("Active"),
                    "created_at": st.column_config.TextColumn("Created", disabled=True),
                },
                disabled=["id", "created_at"],
                key="places_editor",
            )

            if st.button("💾 Save place changes", use_container_width=True):
                changed = 0
                for i in range(len(edited_places)):
                    n = edited_places.iloc[i]
                    o = places_df.iloc[i]
                    pid = str(n["id"])

                    updates = {}
                    if clean_text(n["name"]) != clean_text(o["name"]):
                        updates["name"] = clean_text(n["name"])
                    if bool(n["is_active"]) != bool(o["is_active"]):
                        updates["is_active"] = bool(n["is_active"])

                    if updates:
                        update_place(pid, updates)
                        changed += 1

                st.success(f"Updated {changed} place(s).")
                st.rerun()
